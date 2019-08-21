# -*- coding:UTF-8 -*-
# 读取case文件
import xlrd
import os
from openpyxl import load_workbook
import platform

class request_xl(object):
    def __init__(self, xl_FileName, log):
        self.xl_FileName = xl_FileName
        self.data_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if 'Windows' in platform.system():
            self.data = xlrd.open_workbook(r"%s\configure\%s" % (self.data_path, self.xl_FileName))
            self.log = log
            self.log.info("xlsx path === %s\configure\%s" % (self.data_path, self.xl_FileName))
        else:
            self.data = xlrd.open_workbook(r"%s/configure/%s" % (self.data_path, self.xl_FileName))
            self.log = log
            self.log.info("xlsx path === %s/configure/%s" % (self.data_path, self.xl_FileName))

    def get_xl(self, sheetName, ApiPurpose):
        try:
            # 读取swagger表格的每一个sheet页
            for sheet_name in self.data.sheet_names():
                if sheet_name == sheetName:
                    # 获取的sheet 名称
                    table = self.data.sheet_by_name(sheet_name)
                    # 获取该sheet的行数
                    nrows = table.nrows
                    # 获取该sheet的列数
                    # ncols = table.ncols
                    for i in range(1, nrows):
                        # 使用行列索引，第1列为API Purpose，注：从0开始算
                        API_Purpose = table.row(i)[1].value
                        if API_Purpose == ApiPurpose:
                            self.i = i
                            case = table.row_values(i)   #取出当前一行所有数据，返回为list
                            # self.log.info('test case ========%s' % case)
                            return case
        except Exception as e:
            self.log.error("Get xlsx error ====== %s" % e)



    def update_xl(self, caseFileName, sheetName, ApiPurpose, values):
        try:
            self.log.info("Update excel")
            self.get_xl(sheetName, ApiPurpose)
            # -*- coding: utf-8 -*-
            if 'Windows' in platform.system():
                file_home = r"%s\configure\%s" % (self.data_path, caseFileName)
            else:
                file_home = r"%s/configure/%s" % (self.data_path, caseFileName)
            wb = load_workbook(filename=file_home)
            ws = wb[sheetName]
            rowNum = self.i + 1
            ws['G%s' % rowNum] = values
            wb.save(file_home)
            self.log.info('Update (%s, %s, %s)' % (self.i, 6, values))
        except Exception as e:
            self.log.error("Set xlsx error ====== %s" % e)

